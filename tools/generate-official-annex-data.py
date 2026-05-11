from pathlib import Path
import json
import re
from openpyxl import load_workbook

ROOT = Path(".cache/official_specs/v3_1/2- Annexes_v3.1")
OUT = Path("src/data/officialAnnexData.js")

BT_IDS = {
    "BT-1", "BT-2", "BT-3", "BT-5", "BT-8", "BT-9", "BT-21", "BT-22", "BT-25",
    "BT-30", "BT-31", "BT-40", "BT-47", "BT-48", "BT-55", "BT-63", "BT-72",
    "BT-75", "BT-76", "BT-77", "BT-78", "BT-79", "BT-80", "BT-109", "BT-110",
    "BT-116", "BT-117", "BT-118", "BT-119", "BT-120", "BT-129", "BT-146",
    "BT-147", "BT-148", "BT-153"
}

REPORT_IDS = {
    "TB-1", "TT-1", "TT-2", "TG-1", "TT-3", "TT-4", "TG-3", "TT-8", "TT-7",
    "TT-9", "TT-10", "TG-4", "TT-11", "TT-12", "TG-5", "TT-13", "TT-14",
    "TT-15", "TG-6", "TT-16", "TT-17", "TT-18", "TT-19", "TG-12", "TT-36",
    "TT-37", "TT-38", "TT-39", "TT-40", "TT-41", "TT-42"
}

ANN_IDS = {
    "DG-5-0", "DG-5", "DT-5-2", "DT-5-3", "DT-5-3-1", "DT-5-4", "DT-5-5",
    "DG-6", "DT-6-1", "DT-6-2", "DT-6-3", "DT-6-4"
}


def clean(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def xlsx(pattern):
    return next(ROOT.glob(pattern))


def invoice_rows():
    wb = load_workbook(xlsx("*Annexe 1*.xlsx"), read_only=True, data_only=True)
    rows = {}
    for sheet, syntax in [("FE - Flux 1 - UBL", "UBL"), ("FE - Flux 1 - CII", "CII")]:
        ws = wb[sheet]
        for idx, row in enumerate(ws.iter_rows(min_row=6, values_only=True), start=6):
            ident = clean(row[0])
            if ident not in BT_IDS:
                continue
            entry = rows.setdefault(ident, {
                "id": ident,
                "label": clean(row[2]) or clean(row[3]) or clean(row[4]),
                "cardinality": clean(row[1]),
                "type": clean(row[8]),
                "length": clean(row[9]),
                "values": clean(row[10]),
                "rules": clean(row[15]),
                "definition": clean(row[12]),
                "trajectory": clean(row[14]),
                "base": clean(row[17]),
                "full": clean(row[18]),
                "ubl": "",
                "cii": "",
                "annex": "Annexe 1 - Flux 1 e-invoicing",
                "row": idx
            })
            path = clean(row[6])
            path2 = clean(row[7])
            full_path = f"{path} {path2}".strip()
            if syntax == "UBL":
                entry["ubl"] = full_path
            else:
                entry["cii"] = full_path
    return list(rows.values())


def report_rows():
    wb = load_workbook(xlsx("*Annexe 6*.xlsx"), read_only=True, data_only=True)
    ws = wb["E-REPORTING - Flux 10"]
    out = []
    for idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        ident = clean(row[0])
        if ident not in REPORT_IDS:
            continue
        out.append({
            "id": ident,
            "label": " / ".join(filter(None, [clean(row[2]), clean(row[3]), clean(row[4]), clean(row[5]), clean(row[6])])),
            "cardinality": clean(row[1]),
            "path": clean(row[8]),
            "type": clean(row[9]),
            "length": clean(row[10]),
            "definition": clean(row[11]),
            "values": clean(row[12]),
            "rules": clean(row[13]),
            "b2bInternational": clean(row[14]),
            "b2c": clean(row[15]),
            "subflow": clean(row[16]),
            "annex": "Annexe 6 - Flux 10 e-reporting",
            "row": idx
        })
    corr = wb["E-REPORTING - Correspondance"]
    correspondance = []
    for idx, row in enumerate(corr.iter_rows(min_row=5, values_only=True), start=5):
        if clean(row[1]) or clean(row[2]):
            correspondance.append({
                "cadre": clean(row[1]),
                "categorie": clean(row[2]),
                "annex": "Annexe 6 - Correspondance cadre/categorie",
                "row": idx
            })
    return out, correspondance[:32]


def status_rows():
    wb = load_workbook(xlsx("*Annexe 2*.xlsx"), read_only=True, data_only=True)
    ws = wb["Statuts"]
    out = []
    groups = [
        ("facture_flux2", 1, 2),
        ("donnees_obligatoires_flux1", 4, 5),
        ("ereporting_in_flux10", 7, 8),
        ("ereporting_re_flux10", 10, 11),
    ]
    for idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        for group, code_col, label_col in groups:
            code = clean(row[code_col])
            label = clean(row[label_col])
            if code or label:
                out.append({
                    "id": code,
                    "label": label,
                    "group": group,
                    "annex": "Annexe 2 - Statuts cycle de vie",
                    "row": idx
                })
    return out


def annuaire_rows():
    wb = load_workbook(xlsx("*Annexe 3*.xlsx"), read_only=True, data_only=True)
    out = []
    for sheet in ["FE - F13 (Actualisation)", "FE - F14 (Consultation)"]:
        ws = wb[sheet]
        for idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
            ident = clean(row[0])
            if ident not in ANN_IDS:
                continue
            out.append({
                "id": ident,
                "label": " / ".join(filter(None, [clean(row[2]), clean(row[3]), clean(row[4]), clean(row[5])])),
                "cardinality": clean(row[1]),
                "path": clean(row[8]),
                "type": clean(row[9]),
                "length": clean(row[10]),
                "values": clean(row[11]),
                "annex": f"Annexe 3 - {sheet}",
                "row": idx
            })
    return out


def rules_rows():
    wb = load_workbook(xlsx("*Annexe 7*.xlsx"), read_only=True, data_only=True)
    ws = wb["Règles de gestion"]
    out = []
    for idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        rule = clean(row[1])
        if not rule:
            continue
        if rule in {"G1.01", "G1.02", "G1.05", "G1.07", "G1.09", "G1.10", "G6.08", "G8.01", "G8.02", "G8.05"}:
            out.append({
                "id": rule,
                "title": clean(row[0]),
                "label": clean(row[2]),
                "f1": clean(row[3]),
                "f6": clean(row[4]),
                "f10": clean(row[5]),
                "f13": clean(row[6]),
                "f14": clean(row[7]),
                "annex": "Annexe 7 - Regles de gestion",
                "row": idx
            })
    return out


report, correspondance = report_rows()
data = {
    "invoice": invoice_rows(),
    "reporting": report,
    "reportingCorrespondance": correspondance,
    "statuses": status_rows(),
    "annuaire": annuaire_rows(),
    "rules": rules_rows(),
    "generatedFrom": "Specifications externes v3.1 ZIP officiel impots.gouv.fr, annexes xlsx"
}

OUT.write_text(
    "export const officialAnnexData = " + json.dumps(data, ensure_ascii=False, indent=2) + ";\n",
    encoding="utf-8"
)
print(f"wrote {OUT} with {len(data['invoice'])} invoice rows, {len(data['reporting'])} reporting rows")
